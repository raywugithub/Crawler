from datetime import datetime
import re
import time
from bs4 import BeautifulSoup
from numpy import number
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import math
from openpyxl import load_workbook
import pandas as pd

# Browser settings
chrome_options = Options()
chrome_options.add_argument('--incognito')
chrome_options.add_argument(
    'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.63 Safari/537.36')
# 今天講個特別的，我們可以不讓瀏覽器執行在前景，而是在背景執行（不讓我們肉眼看得見）
# 如以下宣告 options
chrome_options.add_argument('--headless')

chrome_options.add_experimental_option(
    "excludeSwitches", ['enable-automation', 'enable-loggin'])


print('')
print('')
print('===================================================================================================')
# ===================================================================================================
# ===================================================================================================
# 三大法人-區分各期貨契約
browser = webdriver.Chrome(
    chrome_options=chrome_options, executable_path="chromedriver.exe")

browser.get('https://www.taifex.com.tw/cht/3/futContractsDateExcel')
soup = BeautifulSoup(browser.page_source, 'lxml')
futContractsDateExcel_today = soup.find('p').span.find_next('span').string
# print('三大法人-區分各期貨契約 : ', futContractsDateExcel_today)

# 期貨每日交易行情查詢
browser.get('https://www.taifex.com.tw/cht/3/futDailyMarketExcel')
soup = BeautifulSoup(browser.page_source, 'lxml')
futDailyMarketExcel_today = soup.find('p').string
# print('期貨每日交易行情查詢 : ', futDailyMarketExcel_today)

# 三大法人買賣金額統計表
browser.get('https://www.twse.com.tw/fund/BFI82U?response=html')
soup = BeautifulSoup(browser.page_source, 'lxml')
fund_today = soup.find('th').div.string[:-12]
# print('三大法人買賣金額統計表 : ', fund_today)

# 每5秒委託成交統計
browser.get('https://www.twse.com.tw/exchangeReport/MI_5MINS?response=html')
soup = BeautifulSoup(browser.page_source, 'lxml')
mi_5mins_today = soup.find('th').div.string
mi_5mins_today = mi_5mins_today[:-9]
# print('加權指數 每5秒委託成交統計 : ', mi_5mins_today)

# 每5秒指數統計
# browser.get('https://www.twse.com.tw/exchangeReport/MI_5MINS_INDEX?response=html')
# soup = BeautifulSoup(browser.page_source, 'lxml')
# mi_5mins_index_today = soup.find('th').div.string
# mi_5mins_index_today = mi_5mins_index_today[:-7]
# print('加權指數 每5秒指數統計 : ', mi_5mins_index_today)

# 期貨大額交易人未沖銷部位結構表
browser.get('https://www.taifex.com.tw/cht/3/largeTraderFutQryTbl')
soup = BeautifulSoup(browser.page_source, 'lxml')
largeTraderFutQryTbl_today = soup.find('p').string
# print('期貨大額交易人未沖銷部位結構表 : ', largeTraderFutQryTbl_today)

# Close the browser
browser.close()

# weekday() 方法則以整數形式返回一週中的某一天，其中週一的索引為 0，週日為 6
weekday = datetime.today().weekday()
# ===================================================================================================
# ===================================================================================================

print('')
print('')
print('===================================================================================================')
# ===================================================================================================
# ===================================================================================================
# 三大法人-區分各期貨契約
table = pd.read_html(
    'https://www.taifex.com.tw/cht/3/futContractsDateExcel')

df_table = table[1]

# --------------------
# 臺股期貨
# 未平倉餘額 / 多空淨額 / 口數
# 外資
# --------------------
foreign_feature_open_position = df_table[13][5]
foreign_feature_long_open_position = df_table[9][5]
foreign_feature_short_open_position = df_table[11][5]
try:
    foreign_feature_open_position = int(foreign_feature_open_position)
    foreign_feature_long_open_position = int(
        foreign_feature_long_open_position)
    foreign_feature_short_open_position = int(
        foreign_feature_short_open_position)
    print(futContractsDateExcel_today, ' 臺股期貨->外資->未平倉餘額->多空淨額->口數 : ',
          foreign_feature_open_position)
except:
    print('三大法人-區分各期貨契約 ... data are not ready yet!!!')

# --------------------
# 小型臺指期貨
# 未平倉餘額 / 多空淨額 / 口數
# 自營商 + 投信 + 外資
# --------------------
try:
    mtx_three_foreign_feature_open_position = int(df_table[13][12])
    mtx_three_foreign_feature_open_position += int(df_table[13][13])
    mtx_three_foreign_feature_open_position += int(df_table[13][14])
    # print(futContractsDateExcel_today, ' 小型臺指期貨->三大法人->未平倉餘額->多空淨額->加總口數 : ',
    #      mtx_three_foreign_feature_open_position)
except:
    print('小型臺指期貨 ... 未平倉餘額 ... data are not ready yet!!!')
# ===================================================================================================
# ===================================================================================================
# 期貨每日交易行情查詢 / 臺股期貨  ( TX ) 行情表
table = pd.read_html('https://www.taifex.com.tw/cht/3/futDailyMarketExcel')

df_table = table[0]

# --------------------
# 遠月合約
# --------------------
tx_next_month_contract = df_table[1][5]
try:
    tx_next_month_close_price = int(df_table[5][5])
    tx_next_month_open_position = int(df_table[12][5])
    # print(futDailyMarketExcel_today, ' 臺股期貨->遠月合約 : ',
    #      tx_next_month_contract, ' 最後成交價: ', tx_next_month_close_price, ' 未沖銷契約量 : ',
    #      tx_next_month_open_position)
except:
    print('期貨每日交易行情查詢 / 臺股期貨  ( TX ) 行情表 ... data are not ready yet!!!')

# --------------------
# 期貨每日交易行情查詢 / 小型臺指  ( MTX ) 行情表
# --------------------
table = pd.read_html(
    'https://www.taifex.com.tw/cht/3/futDailyMarketExcel?commodity_id=MTX')

df_table = table[0]

# --------------------
# 所有合約 未沖銷契約量
# --------------------
try:
    # if weekday == 2:
    #    mtx_total_open_position = int(df_table[12][12])
    # else:
    #    mtx_total_open_position = int(df_table[12][11])
    mtx_total_open_position = int(df_table[12].dropna().to_list()[-1])
    # print(futDailyMarketExcel_today,
    #      ' 小型臺指->所有合約 未沖銷契約量 : ', mtx_total_open_position)
except:
    print('期貨每日交易行情查詢 / 小型臺指  ( MTX ) 行情表 ... data are not ready yet!!!')

# 散戶多空力道
try:
    mtx_small_open_position = (
        (-1)*mtx_three_foreign_feature_open_position)/mtx_total_open_position
    mtx_small_open_position = round(mtx_small_open_position, 2)
    #print(futDailyMarketExcel_today, ' 散戶多空力道 : ', mtx_small_open_position)
except:
    print('散戶多空力道 ... data are not ready yet!!!')
# ===================================================================================================
# ===================================================================================================
print('===================================================================================================')
# ===================================================================================================
# ===================================================================================================
# 三大法人買賣金額統計表
table = pd.read_html('https://www.twse.com.tw/fund/BFI82U?response=html')

df_table = table[0]
numpy_table = df_table.to_numpy()

# --------------------
# 外資及陸資 / 買賣差額
# --------------------
foreign_fund_trade_number = numpy_table[3][3]
foreign_fund_trade_number = foreign_fund_trade_number / 100000000
foreign_fund_trade_number = round(foreign_fund_trade_number, 2)
print(fund_today, ' 三大法人買賣金額->外資及陸資 : ', foreign_fund_trade_number, ' 億')

# --------------------
# 加權股價指數
# 市場成交資訊
# --------------------
table = pd.read_html(
    'https://www.twse.com.tw/exchangeReport/FMTQIK?response=html')
df_table = table[0]
numpy_table = df_table.to_numpy()
twse_today_diff = numpy_table[-1][-1]
twse_today_diff = round(twse_today_diff, 2)
inside_fund_trade_number = twse_today_diff - foreign_fund_trade_number
inside_fund_trade_number = round(inside_fund_trade_number, 2)
print(fund_today, ' 三大法人買賣金額->內資 : ', inside_fund_trade_number, ' 億')

# --------------------
# 期貨大額交易人未沖銷部位結構表
# --------------------
table = pd.read_html(
    'https://www.taifex.com.tw/cht/3/largeTraderFutQryTbl')
df_table = table[1]
numpy_table = df_table.to_numpy()
all_feature_top_ten_long_open_position = numpy_table[2][4][:6]
all_feature_top_ten_long_open_position = all_feature_top_ten_long_open_position.replace(
    ',', '')
all_feature_top_ten_long_open_position = int(
    all_feature_top_ten_long_open_position)

all_feature_top_ten_short_open_position = numpy_table[2][8][:6]
all_feature_top_ten_short_open_position = all_feature_top_ten_short_open_position.replace(
    ',', '')
all_feature_top_ten_short_open_position = int(
    all_feature_top_ten_short_open_position)

# print(largeTraderFutQryTbl_today, ' 期貨大額交易人未沖銷部位->多方->口數 : ',
#      all_feature_top_ten_long_open_position)
# print(largeTraderFutQryTbl_today, ' 期貨大額交易人未沖銷部位->空方->口數 : ',
#      all_feature_top_ten_short_open_position)
# ===================================================================================================
# ===================================================================================================
print('===================================================================================================')
# ===================================================================================================
# ===================================================================================================
# 讀取 Excel 檔案
wb = load_workbook(
    'everyday_ver2.xlsx')


# 透過名稱取得工作表
sheet = wb['三大法人買賣金額']

if sheet.cell(row=2, column=1).value != fund_today:
    sheet.insert_rows(2)
    sheet.cell(row=2, column=1, value=fund_today)
    sheet.cell(row=2, column=2, value=foreign_fund_trade_number)  # 外資
    sheet.cell(row=2, column=3, value=inside_fund_trade_number)  # 內資

    # 儲存 Excel 檔案
    wb.save('everyday_ver2.xlsx')
    print(fund_today, ' 三大法人買賣金額 ... update !!!')
else:
    print(fund_today, ' 三大法人買賣金額 ... already up to date')


# 透過名稱取得工作表
sheet = wb['散戶多空力道']
if sheet.cell(row=2, column=1).value != futDailyMarketExcel_today:
    sheet.insert_rows(2)
    sheet.cell(row=2, column=1, value=futDailyMarketExcel_today)
    try:
        sheet.cell(row=2, column=2,
                   value=mtx_small_open_position)

        # 儲存 Excel 檔案
        wb.save('everyday_ver2.xlsx')
        print(fund_today, ' 散戶多空力道 ... update !!!')
        print('\t', sheet.cell(row=1, column=2).value, ' : ', sheet.cell(
            row=3, column=2).value, '->', sheet.cell(row=2, column=2).value)
    except:
        print('散戶多空力道 ... data are not ready yet!!!')
else:
    print(futDailyMarketExcel_today, ' 散戶多空力道 ... already up to date')
    print('\t', sheet.cell(row=1, column=2).value, ' : ', sheet.cell(
        row=3, column=2).value, '->', sheet.cell(row=2, column=2).value)


# 透過名稱取得工作表
sheet = wb['台指期換倉成本計算']
if sheet.cell(row=2, column=1).value != futDailyMarketExcel_today:
    sheet.insert_rows(2)
    sheet.cell(row=2, column=1, value=futDailyMarketExcel_today)
    sheet.cell(row=2, column=2, value=tx_next_month_contract)
    sheet.cell(row=2, column=3, value=tx_next_month_close_price)
    sheet.cell(row=2, column=4, value=tx_next_month_open_position)

    try:
        sheet.cell(row=2, column=5, value=(sheet.cell(row=2, column=4).value -
                   sheet.cell(row=3, column=4).value) * sheet.cell(row=2, column=3).value)
    except:
        sheet.cell(row=2, column=5, value=sheet.cell(
            row=2, column=4).value * sheet.cell(row=2, column=3).value)
    total_rows = 0
    for n in range(30):
        if sheet.cell(row=n+2, column=5).value == None:
            total_rows = n
            break
    total_average_price = 0
    for m in range(total_rows):
        total_average_price += sheet.cell(row=m+2, column=5).value
    total_average_price /= sheet.cell(row=2, column=4).value
    sheet.cell(row=2, column=6, value=int(total_average_price))

    # 儲存 Excel 檔案
    wb.save('everyday_ver2.xlsx')
    print(futDailyMarketExcel_today, ' 台指期換倉成本 ... update !!!')
else:
    print(futDailyMarketExcel_today, ' 台指期換倉成本 ... already up to date')


# 透過名稱取得工作表
sheet = wb['大盤多空點位']
if sheet.cell(row=2, column=1).value != mi_5mins_today:
    sheet.insert_rows(2)
    # 每5秒委託成交統計
    # https://www.twse.com.tw/exchangeReport/MI_5MINS?response=html
    table_mi_5mins = pd.read_html(
        "https://www.twse.com.tw/exchangeReport/MI_5MINS?response=html")
    df_table_mi_5mins = table_mi_5mins[0]
    numpy_table_mi_5mins = df_table_mi_5mins.to_numpy()

    # 每5秒指數統計
    # https://www.twse.com.tw/exchangeReport/MI_5MINS_INDEX?response=html
    table_mi_5mins_index = pd.read_html(
        "https://www.twse.com.tw/exchangeReport/MI_5MINS_INDEX?response=html")
    df_table_mi_5mins_index = table_mi_5mins_index[0]
    numpy_table_mi_5mins_index = df_table_mi_5mins_index.to_numpy()

    total_cost_money = 0
    try:
        if len(numpy_table_mi_5mins) == len(numpy_table_mi_5mins_index):
            for n in range(len(numpy_table_mi_5mins) - 1):
                total_cost_money += (numpy_table_mi_5mins[n+1][-2] - numpy_table_mi_5mins[n]
                                     [-2]) * numpy_table_mi_5mins_index[n+1][1]
            twse_one_day_cost = total_cost_money / numpy_table_mi_5mins[-1][-2]
            twse_one_day_cost = round(twse_one_day_cost, 2)
            # twse_one_day_cost = 隔日大盤多空點位

            sheet.cell(row=2, column=1, value=mi_5mins_today)
            sheet.cell(row=2, column=2, value=twse_one_day_cost)

            # 儲存 Excel 檔案
            wb.save(
                'everyday_ver2.xlsx')
            print(mi_5mins_today, ' 大盤多空點位 ... update !!!')
    except:
        print('exception')

else:
    print(mi_5mins_today, ' 大盤多空點位 ... already up to date')


# 期貨大額交易人未沖銷部位
sheet = wb['期貨大額交易人未沖銷部位']
if sheet.cell(row=2, column=1).value != largeTraderFutQryTbl_today:
    sheet.insert_rows(2)
    sheet.cell(row=2, column=1, value=largeTraderFutQryTbl_today)
    sheet.cell(row=2, column=2, value=all_feature_top_ten_long_open_position)
    sheet.cell(row=2, column=3, value=all_feature_top_ten_short_open_position)
    sheet.cell(row=2, column=4, value=(all_feature_top_ten_long_open_position -
               sheet.cell(row=3, column=2).value))
    sheet.cell(row=2, column=5, value=(all_feature_top_ten_short_open_position -
               sheet.cell(row=3, column=3).value))
    sheet.cell(row=2, column=6, value=foreign_feature_long_open_position)
    sheet.cell(row=2, column=7, value=foreign_feature_short_open_position)
    sheet.cell(row=2, column=8, value=foreign_feature_long_open_position -
               sheet.cell(row=3, column=6).value)
    sheet.cell(row=2, column=9, value=foreign_feature_short_open_position -
               sheet.cell(row=3, column=7).value)
    sheet.cell(row=2, column=10, value=foreign_feature_open_position)
    sheet.cell(row=2, column=11, value=foreign_feature_open_position -
               sheet.cell(row=3, column=10).value)
    sheet.cell(row=2, column=12, value=sheet.cell(
        row=2, column=4).value - sheet.cell(row=2, column=8).value)
    sheet.cell(row=2, column=13, value=sheet.cell(
        row=2, column=5).value - sheet.cell(row=2, column=9).value)
    sheet.cell(row=2, column=14, value=sheet.cell(
        row=2, column=12).value - sheet.cell(row=2, column=13).value)

    # 儲存 Excel 檔案
    wb.save(
        'everyday_ver2.xlsx')
    print(mi_5mins_today, ' 期貨大額交易人未沖銷部位 ... update !!!')
    print('\t', sheet.cell(row=1, column=11).value,
          ' : ', sheet.cell(row=2, column=11).value)
    print('\t', sheet.cell(row=1, column=14).value,
          ' : ', sheet.cell(row=2, column=14).value)
else:
    print(mi_5mins_today, ' 期貨大額交易人未沖銷部位 ... already up to date')
    print('\t', sheet.cell(row=1, column=11).value,
          ' : ', sheet.cell(row=2, column=11).value)
    print('\t', sheet.cell(row=1, column=14).value,
          ' : ', sheet.cell(row=2, column=14).value)
# ===================================================================================================
# ===================================================================================================
print('===================================================================================================')
print('')
print('')
