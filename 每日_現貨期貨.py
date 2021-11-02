# coding: utf-8
"""
Post the query to Google　Search and get the return results
"""
import re
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import math
from openpyxl import load_workbook
import pandas as pd

# Browser settings
chrome_options = Options()
chrome_options.add_argument('--incognito')
chrome_options.add_argument(
    'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.63 Safari/537.36')
# 今天講個特別的，我們可以不讓瀏覽器執行在前景，而是在背景執行（不讓我們肉眼看得見）
# 如以下宣告 options
chrome_options.add_argument('--headless')

chrome_options.add_experimental_option(
    "excludeSwitches", ['enable-automation', 'enable-loggin'])

browser = webdriver.Chrome(
    chrome_options=chrome_options)


# 台指期外資單日未平倉.py
browser.get('https://www.taifex.com.tw/cht/3/futContractsDateExcel')
# ===================================================================================================
# ===================================================================================================
print('')
print('')
print('===================================================================================================')

# Crawler
soup = BeautifulSoup(browser.page_source, 'lxml')
content = soup.prettify()

div_section = soup.find('div', class_='section')

table = div_section.find('table')

# date of today
tbody = table.find('tbody')
span_right = tbody.tr.td.p.find('span', class_='right')
futContractsDateExcel_today = span_right.string


# 臺指期貨 - 外資->未平倉餘額->多空淨額->口數
# ---------------------------------------------------------------------------------------------------
# Begin
# ---------------------------------------------------------------------------------------------------
tr_12bk = tbody.find('tr', class_='12bk')
for n in range(3):
    tr_12bk = tr_12bk.find_next('tr')  # 臺股期貨

td_in_tr_12bk = tr_12bk.find_all('td')  # 自營商
# T.B.D.
tr_12bk = tr_12bk.find_next('tr')
td_in_tr_12bk = tr_12bk.find_all('td')  # 投信
# T.B.D.
tr_12bk = tr_12bk.find_next('tr')
td_in_tr_12bk = tr_12bk.find_all('td')  # 外資
# 臺股期貨->外資->未平倉餘額->多空淨額->口數
fitx_foreign_feature_open_position = td_in_tr_12bk[-2]
fitx_foreign_feature_open_position = fitx_foreign_feature_open_position.div.font
try:
    fitx_foreign_feature_open_position = fitx_foreign_feature_open_position.string.strip()
    fitx_foreign_feature_open_position = int(
        fitx_foreign_feature_open_position.replace(',', ''))
except:
    print('exception : 臺股期貨->外資->未平倉餘額->多空淨額->口數 ... data are not ready yet!!!')
# fitx_foreign_feature_open_position = 臺指期貨/外資/單日/未平倉
print(futContractsDateExcel_today, ' : ', '臺股期貨/外資/未平倉餘額/多空淨額/口數 : ',
      fitx_foreign_feature_open_position)
# ---------------------------------------------------------------------------------------------------
# End
# ---------------------------------------------------------------------------------------------------

# 臺指期貨 - 外資成本
# ---------------------------------------------------------------------------------------------------
# Begin
# ---------------------------------------------------------------------------------------------------
browser.get('https://www.taifex.com.tw/cht/3/futDailyMarketExcel')
soup = BeautifulSoup(browser.page_source, 'lxml')
tr_next_tr = soup.find('tr').find_next('tr')
table_next_table = tr_next_tr.find('table').find_next('table')
table_tr_next_tr = table_next_table.find('tr').find_next('tr')
second_table_tr_next_tr = table_tr_next_tr.find_next('tr')
next_month = second_table_tr_next_tr.find_all('td', class_='12bk')[1]
next_month = int(next_month.string.strip())
next_month_open_position = second_table_tr_next_tr.find_all(
    'td', class_='12bk')[-5]
next_month_open_position = int(next_month_open_position.string.strip())
next_month_close = second_table_tr_next_tr.find_all('td', class_='12bk')[-6]
next_month_close = int(next_month_close.string.strip())

futDailyMarketExcel_today = tr_next_tr.p.string
# ---------------------------------------------------------------------------------------------------
# End
# ---------------------------------------------------------------------------------------------------


# 小型臺指期貨 - 計算散戶多空
# ---------------------------------------------------------------------------------------------------
# Begin
# ---------------------------------------------------------------------------------------------------
for n in range(6):
    tr_12bk = tr_12bk.find_next('tr')
# 小型臺指期貨
mitx_foreign_feature_open_position = 0
for n in range(3):
    tr_12bk = tr_12bk.find_next('tr')
    td_in_tr_12bk = tr_12bk.find_all('td')  # 自營商 then 投信 then 外資
    # 小型臺指期貨->外資->未平倉餘額->多空淨額->口數
    mitx_foreign_feature_open_position_temp = td_in_tr_12bk[-2]
    mitx_foreign_feature_open_position_temp = mitx_foreign_feature_open_position_temp.div.font
    try:
        mitx_foreign_feature_open_position_temp = mitx_foreign_feature_open_position_temp.string.strip()
        mitx_foreign_feature_open_position_temp = int(
            mitx_foreign_feature_open_position_temp.replace(',', ''))
        mitx_foreign_feature_open_position += mitx_foreign_feature_open_position_temp
    except:
        print('exception : 散戶多空 ... data are not ready yet!!!')
        break
# mitx_foreign_feature_open_position = 小型臺指期貨/三大法人持有多單倉位

# 散戶持有多單的口數
browser.get(
    'https://www.taifex.com.tw/cht/3/futDailyMarketExcel?commodity_id=MTX')
soup = BeautifulSoup(browser.page_source, 'lxml')
tr_next_tr = soup.find('tr').find_next('tr')
table_next_table = tr_next_tr.find('table').find_next('table')
table_last_tr = table_next_table.find_all('tr')[-1]
tr_last_td = table_last_tr.find_all('td')[-5]
mitx_total_feature_open_position = int(tr_last_td.string)  # 小型臺指期貨/未沖銷契約量

futDailyMarketExcel_mitx_today = tr_next_tr.p.string
# 散戶多空力道
mitx_small_feature_open_position = (-1*mitx_foreign_feature_open_position) / \
    mitx_total_feature_open_position
mitx_small_feature_open_position = round(mitx_small_feature_open_position, 2)
# mitx_small_feature_open_position = 散戶多空力道
# ---------------------------------------------------------------------------------------------------
# End
# ---------------------------------------------------------------------------------------------------


print('===================================================================================================')
# ===================================================================================================
# ===================================================================================================
# 外資 / 內資 買賣超
# ---------------------------------------------------------------------------------------------------
# Begin
# ---------------------------------------------------------------------------------------------------
browser.get(
    'https://www.twse.com.tw/fund/BFI82U?response=html')

# Crawler
soup = BeautifulSoup(browser.page_source, 'lxml')
content = soup.prettify()

twse_today = soup.body.div.table.thead.tr.th.div
twse_today = twse_today.string

tr = soup.body.div.table.tbody.find('tr')
tr = tr.find_next('tr')
tr = tr.find_next('tr')
tr = tr.find_next('tr')  # 外資及陸資(不含外資自營商)
td = tr.find('td')
td = td.find_next('td')
td = td.find_next('td')
td = td.find_next('td')
# print(twse_today, '/外資及陸資(不含外資自營商) : ', td.string)
foreign_fund_number = td.string.replace(',', '')
foreign_fund_number = float(foreign_fund_number)
foreign_fund_number = round(foreign_fund_number / 100000000, 2)
print(twse_today, '/外資及陸資(不含外資自營商) : ', foreign_fund_number, '億')


browser.get('https://www.twse.com.tw/exchangeReport/FMTQIK?response=html')
soup = BeautifulSoup(browser.page_source, 'lxml')
soup = soup.find_all('td')
cat_fund_number = soup[-1].string
cat_fund_number = float(cat_fund_number)
cat_fund_number = (cat_fund_number-foreign_fund_number)
print(twse_today, '/內資 : ', cat_fund_number, '億')
# ---------------------------------------------------------------------------------------------------
# End
# ---------------------------------------------------------------------------------------------------

# 大盤多空點位
# 單日大盤所有成本
# ---------------------------------------------------------------------------------------------------
# Begin
# ---------------------------------------------------------------------------------------------------
# 每5秒委託成交統計
browser.get('https://www.twse.com.tw/exchangeReport/MI_5MINS?response=html')
soup = BeautifulSoup(browser.page_source, 'lxml')
table_mi_5mins_today = soup.find('th').div.string
table_mi_5mins_today = table_mi_5mins_today[:-9]

# 每5秒指數統計
browser.get('https://www.twse.com.tw/exchangeReport/MI_5MINS_INDEX?response=html')
soup = BeautifulSoup(browser.page_source, 'lxml')
table_mi_5mins_index_today = soup.find('th').div.string
table_mi_5mins_index_today = table_mi_5mins_index_today[:-7]
# ---------------------------------------------------------------------------------------------------
# End
# ---------------------------------------------------------------------------------------------------
print('===================================================================================================')
# ===================================================================================================
# ===================================================================================================


# Close the browser
browser.close()


# ===================================================================================================
# ===================================================================================================
# 讀取 Excel 檔案
wb = load_workbook('everyday.xlsx')
# 透過名稱取得工作表
sheet = wb['台指期換倉成本計算']

if sheet.cell(row=2, column=1).value != futDailyMarketExcel_today:
    sheet.insert_rows(2)
    sheet.cell(row=2, column=1, value=futDailyMarketExcel_today)
    sheet.cell(row=2, column=2, value=next_month)
    sheet.cell(row=2, column=3, value=next_month_close)
    sheet.cell(row=2, column=4, value=next_month_open_position)
    try:
        sheet.cell(row=2, column=5, value=(sheet.cell(row=2, column=4).value -
                   sheet.cell(row=3, column=4).value) * sheet.cell(row=2, column=3).value)
    except:
        sheet.cell(row=2, column=5, value=sheet.cell(
            row=2, column=4).value * sheet.cell(row=2, column=3).value)
    total_rows = 0
    for n in range(30):
        if sheet.cell(row=n+2, column=5).value == None:
            total_rows = n
            break
    total_average_price = 0
    for m in range(total_rows):
        total_average_price += sheet.cell(row=m+2, column=5).value
    total_average_price /= sheet.cell(row=2, column=4).value
    sheet.cell(row=2, column=6, value=int(total_average_price))

    print(futDailyMarketExcel_today, '台指期 ', next_month,
          ' 外資成本 : ', sheet.cell(row=2, column=6).value)
    # 儲存 Excel 檔案
    wb.save('everyday.xlsx')
else:
    print(futDailyMarketExcel_today, '台指期 ', next_month,
          ' 外資成本 : ', sheet.cell(row=2, column=6).value, '   ---   up to date')


# 透過名稱取得工作表
sheet = wb['散戶多空力道']
if sheet.cell(row=2, column=1).value != futDailyMarketExcel_mitx_today:
    sheet.insert_rows(2)
    sheet.cell(row=2, column=1, value=futDailyMarketExcel_mitx_today)
    sheet.cell(row=2, column=2,
               value=mitx_small_feature_open_position)

    print(futDailyMarketExcel_mitx_today, '散戶多空力道 : ',
          mitx_small_feature_open_position)
    # 儲存 Excel 檔案
    wb.save('everyday.xlsx')
else:
    print(futDailyMarketExcel_mitx_today, '散戶多空力道 : ',
          sheet.cell(row=2, column=2).value, '   ---   up to date')


# 透過名稱取得工作表
sheet = wb['現貨三大法人']
if sheet.cell(row=2, column=1).value != twse_today[:-12]:
    sheet.insert_rows(2)
    sheet.cell(row=2, column=1, value=twse_today[:-12])
    sheet.cell(row=2, column=2, value=foreign_fund_number)
    sheet.cell(row=2, column=3, value=cat_fund_number)

    print(twse_today[:-12], ' 現貨三大法人 updated')
    # 儲存 Excel 檔案
    wb.save('everyday.xlsx')
else:
    print(twse_today[:-12], ' 現貨三大法人', '   ---   up to date')


# 透過名稱取得工作表
sheet = wb['大盤多空點位']
if sheet.cell(row=2, column=1).value != table_mi_5mins_today:
    sheet.insert_rows(2)
    # 每5秒委託成交統計
    # https://www.twse.com.tw/exchangeReport/MI_5MINS?response=html
    table_mi_5mins = pd.read_html(
        "https://www.twse.com.tw/exchangeReport/MI_5MINS?response=html")
    df_table_mi_5mins = table_mi_5mins[0]
    numpy_table_mi_5mins = df_table_mi_5mins.to_numpy()

    # 每5秒指數統計
    # https://www.twse.com.tw/exchangeReport/MI_5MINS_INDEX?response=html
    table_mi_5mins_index = pd.read_html(
        "https://www.twse.com.tw/exchangeReport/MI_5MINS_INDEX?response=html")
    df_table_mi_5mins_index = table_mi_5mins_index[0]
    numpy_table_mi_5mins_index = df_table_mi_5mins_index.to_numpy()

    total_cost_money = 0
    try:
        if len(numpy_table_mi_5mins) == len(numpy_table_mi_5mins_index):
            for n in range(len(numpy_table_mi_5mins) - 1):
                total_cost_money += (numpy_table_mi_5mins[n+1][-2] - numpy_table_mi_5mins[n]
                                     [-2]) * numpy_table_mi_5mins_index[n+1][1]
            twse_one_day_cost = total_cost_money / numpy_table_mi_5mins[-1][-2]
            twse_one_day_cost = round(twse_one_day_cost, 2)
            # twse_one_day_cost = 隔日大盤多空點位

            sheet.cell(row=2, column=1, value=table_mi_5mins_today)
            sheet.cell(row=2, column=2, value=twse_one_day_cost)

            print(table_mi_5mins_today, ' 大盤多空點位 : ', twse_one_day_cost)
            # 儲存 Excel 檔案
            wb.save('everyday.xlsx')
    except:
        print('exception')

else:
    print(table_mi_5mins_today, ' 大盤多空點位', '   ---   up to date')


print('===================================================================================================')
# ===================================================================================================
# ===================================================================================================

print('')
